from openpyxl import load_workbook


def xlsx_parse(file):
    '''
    Парсим xslx файл. Сортировку решил делать по названию листа в файле.
    Программа ищет листы 'client', 'organization'. Собирает оттуда данные.
    Если таких названий нет, программа считает, что файл содержит счета организаций
    Все данные собираются в один словарь.
    '''
    xls = load_workbook(file)
    sheet_names = xls.sheetnames
    client_list = []
    orgs = {}
    bills = {}

    if 'client' in sheet_names:
        ws = xls['client']
        for row in ws.rows:
            for cell in row:
                if cell.value != 'name':
                    client_list.append(cell.value)

        ws_orgs = xls['organization']
        for row in range(2, ws.max_row + 1):
            client_name = ws_orgs.cell(row=row, column=1).value
            org_name = ws_orgs.cell(row=row, column=2).value
            if client_name not in orgs:
                orgs[client_name] = []
            orgs[client_name].append(org_name)

    else:
        ws = xls.active
        for row in range(2, ws.max_row + 1):
            org_name = ws.cell(row=row, column=1).value
            number = int(ws.cell(row=row, column=2).value)
            amount = ws.cell(row=row, column=3).value
            date = ws.cell(row=row, column=4).value
            if org_name not in bills:
                bills[org_name] = []
            bills[org_name].append([number, amount, date])

    return {'client': client_list, 'organization': orgs, 'bills': bills}



